import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
import os
import matplotlib
matplotlib.use('Agg')  
import matplotlib.pyplot as plt 
import mplfinance as mpf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class ZhuDualEngineScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 
        
        # 使用絕對路徑，確保資料夾 100% 生成在腳本同目錄下
        self.base_dir = os.getcwd()
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.folder_name = f"回後買上漲圖表_{self.today_str}"
        self.folder_path = os.path.join(self.base_dir, self.folder_name)
        
        os.makedirs(self.folder_path, exist_ok=True)

    def get_all_tickers(self):
        tickers = []
        print(">> 正在取得全市場股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
        except Exception: pass 

        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
        except Exception: pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def draw_and_save_chart(self, df, ticker, stock_name, yesterday_high):
        """繪製 K 線圖、3條均線與突破線，並存成圖片"""
        chart_df = df.iloc[-60:].copy() 
        
        apds = [
            mpf.make_addplot(chart_df['5MA'], color='darkorange', width=1.0),
            mpf.make_addplot(chart_df['10MA'], color='purple', width=1.2),
            mpf.make_addplot(chart_df['20MA'], color='deeppink', width=1.5)
        ]
        
        hlines_config = dict(
            hlines=[yesterday_high],
            colors=['blue'],
            linestyle='-.',
            linewidths=1.5
        )
        
        pure_code = ticker.split('.')[0]
        clean_stock_name = str(stock_name).replace("/", "").replace("\\", "").strip()
        filename = os.path.join(self.folder_path, f"{pure_code}_{clean_stock_name}.png")
        
        mc = mpf.make_marketcolors(up='r', down='g', edge='inherit', wick='inherit', volume='in')
        s  = mpf.make_mpf_style(marketcolors=mc, gridstyle=':', y_on_right=True)
        
        os.makedirs(self.folder_path, exist_ok=True)
        
        plot_df = chart_df.tail(125)
        mpf.plot(plot_df, 
            type='candle', 
            volume=True, 
            addplot=apds,           
            hlines=hlines_config,   
            style=s,
            savefig=dict(fname=filename, dpi=100, bbox_inches='tight'),
            figratio=(16,9),  
            figscale=1.2,
            tight_layout=True
        )
        plt.close('all')

    def run_scan(self, tickers, min_vol_lots=1000, chunk_size=50):
        results_level_2 = [] 
        results_level_3 = [] 
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描與繪圖中...")
            
            try:
                # 確保 auto_adjust=True 使用還原權值
                df_all = yf.download(" ".join(chunk), period="3mo", group_by='ticker', progress=False, threads=True, auto_adjust=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < 25: continue 
                    
                    df['5MA'] = df['Close'].rolling(window=5).mean()
                    df['10MA'] = df['Close'].rolling(window=10).mean()
                    df['20MA'] = df['Close'].rolling(window=20).mean()
                    df['5VMA'] = df['Volume'].rolling(window=5).mean()
                    
                    today = df.iloc[-1]
                    yesterday = df.iloc[-2]
                    day_before = df.iloc[-3]
                    
                    is_above_20ma = today['Close'] > today['20MA']
                    is_pullback = (yesterday['Close'] < day_before['Close']) or (yesterday['Close'] < yesterday['Open'])
                    is_break_yesterday_high = today['Close'] > yesterday['High']
                    is_red_candle = today['Close'] > today['Open']
                    is_liquid = today['5VMA'] >= (min_vol_lots * 1000)
                    
                    if not (is_above_20ma and is_pullback and is_break_yesterday_high and is_red_candle and is_liquid):
                        continue 

                    is_vol_shrink = yesterday['Volume'] < day_before['Volume']
                    is_vol_burst = today['Volume'] > (yesterday['Volume'] * 1.5)
                    filter_1_passed = is_vol_shrink and is_vol_burst

                    is_ma_aligned = (today['5MA'] > today['10MA']) and (today['10MA'] > today['20MA'])
                    is_20ma_up = today['20MA'] > yesterday['20MA']
                    filter_2_passed = is_ma_aligned and is_20ma_up

                    dist_10ma = abs(yesterday['Low'] - yesterday['10MA']) / yesterday['10MA']
                    dist_20ma = abs(yesterday['Low'] - yesterday['20MA']) / yesterday['20MA']
                    filter_3_passed = (dist_10ma <= 0.03) or (dist_20ma <= 0.03)

                    if filter_1_passed and filter_2_passed:
                        stock_name = self.ticker_names.get(ticker, "未知") 
                        pure_code = ticker.split('.')[0]
                        yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                        
                        support_str = "未踩支撐"
                        if filter_3_passed:
                            support_str = "踩10MA" if dist_10ma <= 0.03 else "踩20MA"
                            
                        vol_ratio = today['Volume'] / yesterday['Volume'] if yesterday['Volume'] > 0 else 0

                        row_data = {
                            '代號': ticker,
                            '股票名稱': stock_name,
                            '今日收盤': round(today['Close'], 2),
                            '昨日高點': round(yesterday['High'], 2),
                            '今昨量倍數': round(vol_ratio, 1),
                            '今日成交量(張)': int(today['Volume'] / 1000),
                            '昨日成交量(張)': int(yesterday['Volume'] / 1000),
                            '均線狀態': "5>10>20 多頭排列",
                            '回測支撐': support_str,
                            '雅虎股市連結': yahoo_link
                        }
                        
                        # 畫圖
                        self.draw_and_save_chart(df, ticker, stock_name, yesterday['High'])
                        
                        results_level_2.append(row_data)
                        
                        if filter_3_passed:
                            results_level_3.append(row_data)
                            
            except Exception as e:
                print(f"\n[發生隱藏錯誤] {e}\n") 
            
            time.sleep(1.2) 

        return pd.DataFrame(results_level_2), pd.DataFrame(results_level_3)

def export_to_excel(df, filepath, title_color):
    if df.empty: return
    df = df.sort_values(by='今昨量倍數', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "掃描標的"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    
    fill_header = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid") 
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")  
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main; cell.border = border_all
            col_name = headers[c_idx - 1]
            
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"; cell.hyperlink = val; cell.font = font_link; cell.alignment = align_center
            elif col_name in ['代號', '股票名稱', '均線狀態', '回測支撐']:
                cell.value = val; cell.alignment = align_center
                if col_name == '回測支撐' and "踩" in str(val):
                    cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FF0000")
            elif col_name in ['今日收盤', '昨日高點', '今昨量倍數']:
                cell.value = val; cell.alignment = align_right
                cell.number_format = '#,##0.00' if col_name != '今昨量倍數' else '#,##0.0'
            elif col_name in ['今日成交量(張)', '昨日成交量(張)']:
                cell.value = val; cell.alignment = align_right; cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結': cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱': ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        elif headers[col[0].column - 1] == '均線狀態': ws.column_dimensions[col_letter].width = 20
        else: ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    wb.save(filepath)

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = ZhuDualEngineScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (朱家泓回後買：雙報表 + 自動繪圖)...\n")
        df_level_2, df_level_3 = scanner.run_scan(stocks, min_vol_lots=1000)
        
        print("\n" + "="*85)
        print(f"🔥 【條件 1+2】量價配合 + 多頭排列 (符合: {len(df_level_2)} 檔)")
        print(f"🎯 【條件 1+2+3】上述條件 + 靠近均線支撐 (符合: {len(df_level_3)} 檔)")
        print(f"📂 所有產出（Excel 與 K線圖）已完美收納至資料夾：\n👉 【 {scanner.folder_path} 】")
        print("="*85 + "\n")
            
        # 🌟 關鍵修正：將 Excel 儲存路徑直接設在 scanner.folder_path 內部
        folder_path = scanner.folder_path
        export_to_excel(df_level_2, os.path.join(folder_path, f"回後買上漲_標準版(1+2)_{scanner.today_str}.xlsx"), title_color="366092")
        export_to_excel(df_level_3, os.path.join(folder_path, f"回後買上漲_嚴格版(1+2+3)_{scanner.today_str}.xlsx"), title_color="990000")