
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    raw_download = getattr(yf, 'download')
    df = raw_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df




















import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import numpy as np
import requests
import time
import urllib3
import os
import matplotlib
matplotlib.use('Agg')  
import matplotlib.pyplot as plt 
import mplfinance as mpf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class ABCBreakoutScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 
        
        self.base_dir = os.getcwd()
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.folder_name = f"ABC突破切線圖表_{self.today_str}"
        self.folder_path = os.path.join(self.base_dir, self.folder_name)
        
        os.makedirs(self.folder_path, exist_ok=True)

    def get_all_tickers(self):
        tickers = []
        print(">> 正在取得全市場股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
        except Exception: pass 

        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
        except Exception: pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def zigzag_peaks_valleys(self, df, deviation=0.02):
        peaks = []
        valleys = []
        
        highs = df['High'].values
        lows = df['Low'].values
        
        last_extreme_idx = 0
        last_extreme_price = highs[0]
        direction = 1 
        
        for i in range(1, len(df)):
            if direction == 1: 
                if highs[i] > last_extreme_price:
                    last_extreme_price = highs[i]
                    last_extreme_idx = i
                elif lows[i] < last_extreme_price * (1 - deviation):
                    peaks.append((last_extreme_idx, last_extreme_price))
                    direction = -1
                    last_extreme_price = lows[i]
                    last_extreme_idx = i
            else: 
                if lows[i] < last_extreme_price:
                    last_extreme_price = lows[i]
                    last_extreme_idx = i
                elif highs[i] > last_extreme_price * (1 + deviation):
                    valleys.append((last_extreme_idx, last_extreme_price))
                    direction = 1
                    last_extreme_price = highs[i]
                    last_extreme_idx = i
                    
        return peaks, valleys

    def draw_and_save_chart(self, df, ticker, stock_name, peak1_idx, peak1_val, peak2_idx, peak2_val, current_idx):
        start_idx = max(0, peak1_idx - 20)
        chart_df = df.iloc[start_idx:].copy()
        
        rel_p1_idx = peak1_idx - start_idx
        rel_p2_idx = peak2_idx - start_idx
        rel_curr_idx = current_idx - start_idx
        
        m = (peak2_val - peak1_val) / (rel_p2_idx - rel_p1_idx)
        c = peak1_val - (m * rel_p1_idx)
        
        trendline = [np.nan] * len(chart_df)
        for i in range(rel_p1_idx, len(chart_df)):
            trendline[i] = m * i + c
            
        apds = [
            mpf.make_addplot(chart_df['20MA'], color='deeppink', width=1.5),
            mpf.make_addplot(trendline, color='#00AA00', width=2.0, type='line')
        ]
        
        pure_code = ticker.split('.')[0]
        clean_stock_name = str(stock_name).replace("/", "").replace("\\", "").strip()
        filename = os.path.join(self.folder_path, f"{pure_code}_{clean_stock_name}.png")
        
        mc = mpf.make_marketcolors(up='r', down='g', edge='inherit', wick='inherit', volume='in')
        s  = mpf.make_mpf_style(marketcolors=mc, gridstyle=':', y_on_right=True)
        
        os.makedirs(self.folder_path, exist_ok=True)
        
        # 🌟 已修復這裡的字串組合錯誤
        chart_title = f"{pure_code} {clean_stock_name} - ABC Breakout"
        
        plot_df = chart_df.tail(125)
        hank_6m_df = plot_df.tail(125)
        mpf.plot(hank_6m_df, 
            type='candle', 
            volume=True, 
            addplot=apds,           
            style=s,
            savefig=dict(fname=filename, dpi=100, bbox_inches='tight'),
            figratio=(16,9),  
            figscale=1.2,
            tight_layout=True,
            title=chart_title
        )
        plt.close('all')

    def run_scan(self, tickers, deviation_pct=0.03, min_vol_lots=1000, chunk_size=50):
        results = []
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描與繪圖中...")
            
            try:
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True, auto_adjust=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < 60: continue 
                    
                    df['20MA'] = df['Close'].rolling(window=20).mean()
                    df['5VMA'] = df['Volume'].rolling(window=5).mean()
                    
                    peaks, valleys = self.zigzag_peaks_valleys(df, deviation=deviation_pct)
                    
                    if len(peaks) < 2 or len(valleys) < 2: continue
                    
                    peak2_idx, peak2_val = peaks[-1]
                    peak1_idx, peak1_val = peaks[-2]
                    
                    current_idx = len(df) - 1
                    
                    if (current_idx - peak2_idx) > 15: continue
                    if peak2_val >= peak1_val: continue
                    
                    valleys_between = [v for v in valleys if peak1_idx < v[0] < peak2_idx]
                    if not valleys_between: continue
                    valleyA_idx, valleyA_val = valleys_between[0]
                    
                    valleys_after = [v for v in valleys if peak2_idx < v[0] < current_idx]
                    if not valleys_after: continue
                    valleyC_idx, valleyC_val = valleys_after[-1]
                    
                    if valleyC_val >= valleyA_val: continue
                    
                    m = (peak2_val - peak1_val) / (peak2_idx - peak1_idx)
                    c = peak1_val - (m * peak1_idx)
                    
                    trendline_today = m * current_idx + c
                    trendline_yest = m * (current_idx - 1) + c
                    
                    today = df.iloc[-1]
                    yesterday = df.iloc[-2]
                    
                    is_red_candle = today['Close'] > today['Open']
                    is_breakout_today = today['Close'] > trendline_today
                    is_below_yest = yesterday['Close'] <= trendline_yest
                    is_above_20ma = today['Close'] > today['20MA']
                    
                    avg_vol_5d = today['5VMA']
                    is_liquid = avg_vol_5d >= (min_vol_lots * 1000)
                    is_vol_up = today['Volume'] > yesterday['Volume']
                    
                    if is_red_candle and is_breakout_today and is_below_yest and is_above_20ma and is_liquid and is_vol_up:
                        
                        stock_name = self.ticker_names.get(ticker, "未知") 
                        pure_code = ticker.split('.')[0]
                        yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                        vol_ratio = today['Volume'] / yesterday['Volume'] if yesterday['Volume'] > 0 else 0
                        
                        self.draw_and_save_chart(df, ticker, stock_name, peak1_idx, peak1_val, peak2_idx, peak2_val, current_idx)
                        
                        results.append({
                            '代號': ticker,
                            '股票名稱': stock_name,
                            '今日收盤': round(today['Close'], 2),
                            '切線壓力價': round(trendline_today, 2),
                            '突破幅度': f"{round((today['Close'] - trendline_today)/trendline_today * 100, 2)}%",
                            '今日成交量(張)': int(today['Volume'] / 1000),
                            '今昨量倍數': round(vol_ratio, 1),
                            '雅虎股市連結': yahoo_link
                        })
                        
            except Exception as e:
                pass
            
            time.sleep(1.2) 

        return pd.DataFrame(results)

def export_to_excel(df, filepath):
    if df.empty: return
    
    df = df.sort_values(by='今昨量倍數', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ABC突破精選"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    font_highlight = Font(name="微軟正黑體", size=11, bold=True, color="FF0000") 
    
    fill_header = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    fill_zebra = PatternFill(start_color="F0FAFA", end_color="F0FAFA", fill_type="solid")  
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main; cell.border = border_all
            
            col_name = headers[c_idx - 1]
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = align_center
            elif col_name in ['代號', '股票名稱']:
                cell.value = val; cell.alignment = align_center
            elif col_name == '突破幅度':
                cell.value = val; cell.alignment = align_center; cell.font = font_highlight 
            elif col_name in ['今日收盤', '切線壓力價', '今昨量倍數']:
                cell.value = val; cell.alignment = align_right
                cell.number_format = '#,##0.00' if col_name != '今昨量倍數' else '#,##0.0'
            elif col_name == '今日成交量(張)':
                cell.value = val; cell.alignment = align_right; cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結': cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱': ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        else: ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    wb.save(filepath)

if __name__ == "__main__":
    scanner = ABCBreakoutScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (突破 ABC 修正下降切線 - ZigZag AI 降噪版)...\n")
        
        final_df = scanner.run_scan(stocks, deviation_pct=0.03, min_vol_lots=1000)
        
        print("\n" + "="*95)
        if not final_df.empty:
            print(f"📈 掃描完成！共發現 {len(final_df)} 檔突破【ABC下降切線】的表態股。")
            print(f"📂 所有產出（藍綠 Excel 與 切線K線圖）已完美收納至資料夾：\n👉 【 {scanner.folder_path} 】")
            print(f"   (圖表說明 -> 粉紅線: 20MA | 綠實線: 電腦自動畫出的下降切線壓力)")
            
            excel_filepath = os.path.join(scanner.folder_path, f"ABC突破切線精選_{scanner.today_str}.xlsx")
            export_to_excel(final_df, excel_filepath)
        else:
            print("❌ 今日無符合「突破 ABC 修正下降切線」的標的。")
        print("="*95 + "\n")