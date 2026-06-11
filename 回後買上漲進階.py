
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    # 🌟 使用 getattr 避開字串 replace 關鍵字，完美斷開無限遞迴死結
    raw_download = getattr(yf, 'download')
    df = raw_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df














import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class TaiwanStockScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 

    def get_all_tickers(self):
        """從證交所與櫃買中心取得【上市+上櫃】所有股票代號與名稱"""
        tickers = []
        
        print(">> 正在取得【上市】股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            data = response.json()
            for item in data:
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
            print(f"   成功取得上市標的。")
        except Exception as e:
            pass 

        print(">> 正在取得【上櫃】股票清單...")
        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            data = response.json()
            for item in data:
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
            print(f"   成功取得上櫃標的。")
        except Exception as e:
            pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def run_scan(self, tickers, min_vol_lots=1000, chunk_size=50):
        """一次掃描，產出「1+2」與「1+2+3」兩組結果"""
        results_level_2 = [] # 存放 1+2 條件的結果
        results_level_3 = [] # 存放 1+2+3 條件的結果
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描中...")
            
            try:
                # 抓取3個月資料
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < 25: continue 
                    
                    # 1. 計算所有技術指標
                    df['5MA'] = df['Close'].rolling(window=5).mean()
                    df['10MA'] = df['Close'].rolling(window=10).mean()
                    df['20MA'] = df['Close'].rolling(window=20).mean()
                    df['5VMA'] = df['Volume'].rolling(window=5).mean()
                    
                    today = df.iloc[-1]
                    yesterday = df.iloc[-2]
                    day_before = df.iloc[-3]
                    
                    # =============== 【基礎條件】朱家泓回後買上漲 ===============
                    is_above_20ma = today['Close'] > today['20MA']
                    is_pullback = (yesterday['Close'] < day_before['Close']) or (yesterday['Close'] < yesterday['Open'])
                    is_break_yesterday_high = today['Close'] > yesterday['High']
                    is_red_candle = today['Close'] > today['Open']
                    is_liquid = today['5VMA'] >= (min_vol_lots * 1000)
                    
                    base_passed = is_above_20ma and is_pullback and is_break_yesterday_high and is_red_candle and is_liquid
                    
                    if not base_passed:
                        continue # 基礎條件沒過，直接換下一檔

                    # =============== 【濾網 1】量價配合 ===============
                    # 昨量縮：昨天成交量 < 前天成交量
                    is_vol_shrink = yesterday['Volume'] < day_before['Volume']
                    # 今出量：今天成交量 > 昨天成交量 1.5 倍
                    is_vol_burst = today['Volume'] > (yesterday['Volume'] * 1.5)
                    filter_1_passed = is_vol_shrink and is_vol_burst

                    # =============== 【濾網 2】多頭排列 ===============
                    # 均線多頭排列：5MA > 10MA > 20MA
                    is_ma_aligned = (today['5MA'] > today['10MA']) and (today['10MA'] > today['20MA'])
                    # 月線斜率向上
                    is_20ma_up = today['20MA'] > yesterday['20MA']
                    filter_2_passed = is_ma_aligned and is_20ma_up

                    # =============== 【濾網 3】靠近支撐發動 ===============
                    # 昨天回檔的最低點，距離 10MA 或 20MA 在 3% 以內
                    dist_10ma = abs(yesterday['Low'] - yesterday['10MA']) / yesterday['10MA']
                    dist_20ma = abs(yesterday['Low'] - yesterday['20MA']) / yesterday['20MA']
                    filter_3_passed = (dist_10ma <= 0.03) or (dist_20ma <= 0.03)

                    # 整理輸出資料
                    if filter_1_passed and filter_2_passed:
                        stock_name = self.ticker_names.get(ticker, "未知") 
                        pure_code = ticker.split('.')[0]
                        yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                        
                        # 支撐判斷文字
                        support_str = "未踩支撐"
                        if filter_3_passed:
                            support_str = "踩10MA" if dist_10ma <= 0.03 else "踩20MA"
                            
                        # 計算今昨量比
                        vol_ratio = today['Volume'] / yesterday['Volume'] if yesterday['Volume'] > 0 else 0

                        row_data = {
                            '代號': ticker,
                            '股票名稱': stock_name,
                            '今日收盤': round(today['Close'], 2),
                            '昨日高點': round(yesterday['High'], 2),
                            '今昨量倍數': round(vol_ratio, 1),
                            '今日成交量(張)': int(today['Volume'] / 1000),
                            '昨日成交量(張)': int(yesterday['Volume'] / 1000),
                            '均線狀態': "5>10>20 多頭排列",
                            '回測支撐': support_str,
                            '雅虎股市連結': yahoo_link
                        }
                        
                        # 加入 1+2 清單
                        results_level_2.append(row_data)
                        
                        # 如果連濾網 3 都過關，再加入 1+2+3 清單
                        if filter_3_passed:
                            results_level_3.append(row_data)
                            
            except Exception as e:
                pass 
            
            time.sleep(1.2) 

        return pd.DataFrame(results_level_2), pd.DataFrame(results_level_3)

def export_to_excel(df, filename, title_color):
    """將結果導出為精美樣式的 Excel 表格"""
    if df.empty:
        print(f">> 沒有符合條件的資料，不產生 Excel 檔：{filename}")
        return
    
    # 依據今昨量倍數(爆發力)排序
    df = df.sort_values(by='今昨量倍數', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "掃描標的"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    
    fill_header = PatternFill(start_color=title_color, end_color=title_color, fill_type="solid") 
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")  
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main
            cell.border = border_all
            
            col_name = headers[c_idx - 1]
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = align_center
            elif col_name in ['代號', '股票名稱', '均線狀態', '回測支撐']:
                cell.value = val
                cell.alignment = align_center
                
                # 若為踩支撐，將文字標紅加粗提示
                if col_name == '回測支撐' and "踩" in str(val):
                    cell.font = Font(name="微軟正黑體", size=11, bold=True, color="FF0000")
                    
            elif col_name in ['今日收盤', '昨日高點', '今昨量倍數']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0.00' if col_name != '今昨量倍數' else '#,##0.0'
            elif col_name in ['今日成交量(張)', '昨日成交量(張)']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結':
                cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱':
            ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        elif headers[col[0].column - 1] == '均線狀態':
            ws.column_dimensions[col_letter].width = 20
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    
    wb.save(filename)
    print(f">> ✅ 成功產出 Excel：{filename}")

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = TaiwanStockScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (雙引擎運算中)...\n")
        
        # 執行掃描，同時接住兩組 DataFrame
        df_level_2, df_level_3 = scanner.run_scan(stocks, min_vol_lots=1000)

# 💡 加上日期檔名
        today_str = datetime.now().strftime("%Y%m%d")
        file_l2 = f"回後買上漲_標準版(1+2)_{today_str}.xlsx"
        file_l3 = f"回後買上漲_嚴格版(1+2+3)_{today_str}.xlsx"
            
        # 傳入含有日期的全新檔名
        export_to_excel(df_level_2, filename=file_l2, title_color="366092")
        export_to_excel(df_level_3, filename=file_l3, title_color="990000")
        
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)
        
        print("\n" + "="*85)
        print(f"🔥 【條件 1+2】量價配合 + 多頭排列 (符合: {len(df_level_2)} 檔)")
        print(f"🎯 【條件 1+2+3】上述條件 + 靠近均線支撐 (符合: {len(df_level_3)} 檔)")
        print("="*85 + "\n")
            
        # 輸出第一份：1+2 標準版 (使用藍色表頭)
        export_to_excel(df_level_2, "回後買上漲_標準版(1+2).xlsx", title_color="366092")
        
        # 輸出第二份：1+2+3 嚴格版 (使用深紅色表頭，警示這更具狙擊價值)
        export_to_excel(df_level_3, "回後買上漲_嚴格版(1+2+3).xlsx", title_color="990000")