
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    raw_download = getattr(yf, 'download')
    df = raw_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df




















import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import mplfinance as mpf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class DibaoScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 
        
        # 🌟 完美收納設定：使用絕對路徑，確保資料夾 100% 生成在腳本同目錄下
        self.base_dir = os.getcwd()
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.folder_name = f"帝寶線圖表_{self.today_str}"
        self.folder_path = os.path.join(self.base_dir, self.folder_name)
        
        # 建立資料夾
        os.makedirs(self.folder_path, exist_ok=True)

    def get_all_tickers(self):
        """從證交所與櫃買中心取得【上市+上櫃】所有股票代號與名稱"""
        tickers = []
        print(">> 正在取得全市場股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
        except Exception: pass 

        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
        except Exception: pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def draw_and_save_chart(self, df, ticker, stock_name, gap_down_price, engulfing_price):
        """繪製 K 線圖，強調開低點與吞噬點"""
        chart_df = df.iloc[-125:].copy() # 底部反轉，看近一個半月即可
        
        # 設定關鍵水平線：昨日黑K收盤(開低基準)、昨日黑K開盤(吞噬基準)
        hlines_config = dict(
            hlines=[gap_down_price, engulfing_price],
            colors=['#0088CC', '#FF0000'], # 藍：前收、紅：前高
            linestyle='--',
            linewidths=1.2
        )
        
        # 檔名處理 (避開特殊字元)
        pure_code = ticker.split('.')[0]
        clean_stock_name = str(stock_name).replace("/", "").replace("\\", "").strip()
        filename = os.path.join(self.folder_path, f"{pure_code}_{clean_stock_name}.png")
        
        # 台股紅綠 K 線風格
        mc = mpf.make_marketcolors(up='r', down='g', edge='inherit', wick='inherit', volume='in')
        s  = mpf.make_mpf_style(marketcolors=mc, gridstyle=':', y_on_right=True)
        
        # 雙重確認資料夾
        os.makedirs(self.folder_path, exist_ok=True)
        
        plot_df = chart_df.tail(125)
        hank_6m_df = plot_df.tail(125)
        mpf.plot(hank_6m_df, 
            type='candle', 
            volume=True, 
            hlines=hlines_config,
            style=s,
            savefig=dict(fname=filename, dpi=100, bbox_inches='tight'),
            figratio=(16,9),  
            figscale=1.2,
            tight_layout=True
        )
        plt.close('all')

    def run_scan(self, tickers, lookback_days=20, min_vol_lots=1000, chunk_size=50):
        """執行批次掃描 (入住帝寶線)"""
        results = []
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描與繪圖中...")
            
            try:
                # 🌟 確保 auto_adjust=True 還原權值
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True, auto_adjust=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < (lookback_days + 5): continue 
                    
                    df['5VMA'] = df['Volume'].rolling(window=5).mean()
                    
                    today = df.iloc[-1]
                    yesterday = df.iloc[-2]
                    
                    # 1. 位階判斷：確保出現在波段低點 (昨天的最低價是過去 20 天以來的最低)
                    past_20_days_low = df['Low'].iloc[-(lookback_days+1):-1].min()
                    is_bottom_area = yesterday['Low'] <= past_20_days_low
                    
                    # 2. 型態判斷
                    # 昨日為誘空黑K
                    is_yesterday_black = yesterday['Close'] < yesterday['Open']
                    # 今日跳空開低
                    is_gap_down = today['Open'] < yesterday['Close']
                    # 今日強勢吞噬 (實體紅K完全吃掉昨日黑K實體)
                    is_engulfing = today['Close'] > yesterday['Open']
                    
                    # 3. 量能判斷
                    is_volume_up = today['Volume'] > yesterday['Volume']
                    is_liquid = today['5VMA'] >= (min_vol_lots * 1000)
                    
                    # 綜合判定
                    if is_bottom_area and is_yesterday_black and is_gap_down and is_engulfing and is_volume_up and is_liquid:
                        
                        stock_name = self.ticker_names.get(ticker, "未知") 
                        pure_code = ticker.split('.')[0]
                        yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                        
                        breakout_strength = (today['Close'] - yesterday['Open']) / yesterday['Open']
                        
                        # 呼叫畫圖，傳入昨日收盤(藍線)與昨日開盤(紅線)
                        self.draw_and_save_chart(df, ticker, stock_name, yesterday['Close'], yesterday['Open'])
                        
                        results.append({
                            '代號': ticker,
                            '股票名稱': stock_name,
                            '今日收盤': round(today['Close'], 2),
                            '今日開盤(開低)': round(today['Open'], 2),
                            '昨日收盤(黑K)': round(yesterday['Close'], 2),
                            '昨日開盤(黑K)': round(yesterday['Open'], 2),
                            '吞噬強度': f"{round(breakout_strength * 100, 1)}%",
                            '今日成交量(張)': int(today['Volume'] / 1000),
                            '五日均量(張)': int(today['5VMA'] / 1000),
                            '雅虎股市連結': yahoo_link
                        })
                        
            except Exception as e:
                print(f"\n[發生隱藏錯誤] {e}\n") 
            
            time.sleep(1.2) 

        return pd.DataFrame(results)

def export_to_excel(df, filepath):
    """將結果導出為精美樣式的 Excel 表格"""
    if df.empty: return
    
    df = df.sort_values(by='吞噬強度', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "帝寶線精選"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="000000") # 黑字
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    font_highlight = Font(name="微軟正黑體", size=11, bold=True, color="FF0000") 
    
    fill_header = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid") # 金黃色
    fill_zebra = PatternFill(start_color="FFFDF0", end_color="FFFDF0", fill_type="solid")  
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main
            cell.border = border_all
            
            col_name = headers[c_idx - 1]
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = align_center
            elif col_name in ['代號', '股票名稱']:
                cell.value = val
                cell.alignment = align_center
            elif col_name == '吞噬強度':
                cell.value = val
                cell.alignment = align_center
                cell.font = font_highlight 
            elif col_name in ['今日收盤', '今日開盤(開低)', '昨日收盤(黑K)', '昨日開盤(黑K)']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0.00'
            elif col_name in ['今日成交量(張)', '五日均量(張)']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結':
                cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱':
            ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    
    wb.save(filepath)

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = DibaoScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (入住帝寶線：雙報表 + 自動繪圖)...\n")
        
        final_df = scanner.run_scan(stocks, lookback_days=20, min_vol_lots=1000)
        
        print("\n" + "="*95)
        if not final_df.empty:
            print(f"🏰 掃描完成！共發現 {len(final_df)} 檔【入住帝寶線】標的。")
            print(f"📂 所有產出（金黃 Excel 與 K線圖）已完美收納至資料夾：\n👉 【 {scanner.folder_path} 】")
            print(f"   (圖表虛線說明 -> 藍色: 昨收(跳空開低點) | 紅色: 昨開(完全吞噬點))")
            
            # 🌟 關鍵修正：將 Excel 儲存路徑直接設在 scanner.folder_path 內部
            excel_filepath = os.path.join(scanner.folder_path, f"帝寶線_多頭吞噬精選_{scanner.today_str}.xlsx")
            export_to_excel(final_df, excel_filepath)
        else:
            print("❌ 今日全市場無符合嚴格「入住帝寶線」特徵的標的。耐心等待主力洗盤！")
        print("="*95 + "\n")