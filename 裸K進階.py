
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    # 🌟 使用 getattr 避開字串 replace 關鍵字，完美斷開無限遞迴死結
    raw_download = getattr(yf, 'download')
    df = raw_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df














import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
import os
import matplotlib
matplotlib.use('Agg')  # 解決多執行緒背景畫圖當機
import matplotlib.pyplot as plt 
import mplfinance as mpf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class LaoYuNakedKScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 
        
        # 使用絕對路徑，確保資料夾 100% 生成在腳本同目錄下
        self.base_dir = os.getcwd()
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.folder_name = f"老余裸K圖表_{self.today_str}"
        self.folder_path = os.path.join(self.base_dir, self.folder_name)
        
        os.makedirs(self.folder_path, exist_ok=True)

    def get_all_tickers(self):
        """從證交所與櫃買中心取得【上市+上櫃】所有股票代號與名稱"""
        tickers = []
        print(">> 正在取得全市場股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
        except Exception: pass 

        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
        except Exception: pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def draw_and_save_chart(self, df, ticker, stock_name, entry, stop_loss, target):
        """繪製 K 線圖與關鍵進出場線，並存成圖片"""
        chart_df = df.iloc[-125:].copy() 
        
        # 設定三條關鍵水平線 (綠: 目標壓力, 藍: 今日進場, 紅: 破底停損)
        hlines_config = dict(
            hlines=[target, entry, stop_loss],
            colors=['g', 'b', 'r'],
            linestyle='-.',
            linewidths=1.5
        )
        
        # 過濾股票名稱中的特殊字元，避免存檔失敗
        pure_code = ticker.split('.')[0]
        clean_stock_name = str(stock_name).replace("/", "").replace("\\", "").strip()
        filename = os.path.join(self.folder_path, f"{pure_code}_{clean_stock_name}.png")
        
        mc = mpf.make_marketcolors(up='r', down='g', edge='inherit', wick='inherit', volume='in')
        s  = mpf.make_mpf_style(marketcolors=mc, gridstyle=':', y_on_right=True)
        
        # 繪圖前確認資料夾存在
        os.makedirs(self.folder_path, exist_ok=True)
        
        plot_df = chart_df.tail(125)
        hank_6m_df = plot_df.tail(125)
        mpf.plot(hank_6m_df, 
            type='candle', 
            volume=True, 
            hlines=hlines_config,
            style=s,
            savefig=dict(fname=filename, dpi=100, bbox_inches='tight'),
            figratio=(16,9),  
            figscale=1.2,
            tight_layout=True
        )
        # 正規釋放記憶體
        plt.close('all')

    def run_scan(self, tickers, support_lookback=40, trap_window=5, min_rr_ratio=2.0, min_vol_lots=500, chunk_size=50):
        results = []
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描與繪圖中...")
            
            try:
                # 🌟 確保 auto_adjust=True 還原權值，看見真正的 K 線
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True, auto_adjust=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < (support_lookback + trap_window + 5): continue 
                    
                    history_zone = df.iloc[-(support_lookback + trap_window) : -trap_window]
                    trap_zone = df.iloc[-trap_window : -1]
                    today = df.iloc[-1]
                    
                    support_price = history_zone['Low'].min()      
                    resistance_price = history_zone['High'].max()  
                    trap_low = trap_zone['Low'].min()              
                    entry_price = today['Close']                   
                    
                    is_breakdown = trap_low < support_price
                    is_reclaim = entry_price > support_price
                    is_red_candle = entry_price > today['Open']
                    is_liquid = today['Volume'] >= (min_vol_lots * 1000)
                    
                    if is_breakdown and is_reclaim and is_red_candle and is_liquid:
                        risk = entry_price - trap_low          
                        reward = resistance_price - entry_price 
                        
                        if risk > 0 and reward > 0:
                            rr_ratio = reward / risk
                            
                            if rr_ratio >= min_rr_ratio:
                                stock_name = self.ticker_names.get(ticker, "未知") 
                                pure_code = ticker.split('.')[0]
                                yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                                
                                # 呼叫畫圖函數
                                self.draw_and_save_chart(df, ticker, stock_name, entry_price, trap_low, resistance_price)
                                
                                results.append({
                                    '代號': ticker,
                                    '股票名稱': stock_name,
                                    '賺賠比(RR)': round(rr_ratio, 2),
                                    '進場價(今日收盤)': round(entry_price, 2),
                                    '停損價(破底低點)': round(trap_low, 2),
                                    '目標價(前波壓力)': round(resistance_price, 2),
                                    '風險空間(元)': round(risk, 2),
                                    '今日成交量(張)': int(today['Volume'] / 1000),
                                    '雅虎股市連結': yahoo_link
                                })
                                
            except Exception as e:
                print(f"\n[發生隱藏錯誤] {e}\n") 
            
            time.sleep(1.2) 

        return pd.DataFrame(results)

def export_to_excel(df, filepath):
    if df.empty: return
    df = df.sort_values(by='賺賠比(RR)', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "裸K賺賠比精選"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    font_highlight = Font(name="微軟正黑體", size=11, bold=True, color="008000") 
    
    fill_header = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid") 
    fill_zebra = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")  
    border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header; cell.fill = fill_header; cell.alignment = align_center; cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main; cell.border = border_all
            col_name = headers[c_idx - 1]
            
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"; cell.hyperlink = val; cell.font = font_link; cell.alignment = align_center
            elif col_name in ['代號', '股票名稱']:
                cell.value = val; cell.alignment = align_center
            elif col_name == '賺賠比(RR)':
                cell.value = f"1 : {val}"; cell.alignment = align_center; cell.font = font_highlight 
            elif col_name in ['進場價(今日收盤)', '停損價(破底低點)', '目標價(前波壓力)', '風險空間(元)']:
                cell.value = val; cell.alignment = align_right; cell.number_format = '#,##0.00'
            elif col_name == '今日成交量(張)':
                cell.value = val; cell.alignment = align_right; cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結': cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱': ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        else: ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    wb.save(filepath)

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = LaoYuNakedKScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (老余流：裸K破底翻 + 自動繪製進出場圖表)...\n")
        # 您可以調整 min_rr_ratio 來改變嚴格度 (預設 2.0 代表賺賠比需 > 1:2)
        final_df = scanner.run_scan(stocks, min_rr_ratio=2.0)
        
        print("\n" + "="*90)
        if not final_df.empty:
            print(f"🔥 掃描完成！共發現 {len(final_df)} 檔【高性價比】標的。")
            print(f"📂 所有產出（Excel 與 K線圖）已完美收納至資料夾：\n👉 【 {scanner.folder_path} 】")
            print(f"   (圖表虛線說明 -> 綠色: 目標壓力 | 藍色: 今日進場 | 紅色: 破底停損)")
            
            # 🌟 關鍵修正：將 Excel 儲存路徑直接設在 scanner.folder_path 內部
            excel_filepath = os.path.join(scanner.folder_path, f"老余裸K_賺賠比精選_{scanner.today_str}.xlsx")
            export_to_excel(final_df, excel_filepath)
        else:
            print("❌ 殘酷的市場：今日無符合「賺賠比 >= 2.0」的破底翻標的。不勉強進場！")
        print("="*90 + "\n")