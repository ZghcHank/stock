
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        # 🟢 0秒硬碟解鎖機制：存在當日快取就直接秒讀
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    # 若硬碟沒檔案 (網路防護補償網)，才真正發動網路下載
    df = cached_yf_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df


import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
import os
import matplotlib
matplotlib.use('Agg')  # 解決多執行緒背景畫圖當機
import matplotlib.pyplot as plt
import mplfinance as mpf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class TaiwanStockScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 
        
        # 🌟 完美收納設定：建立以日期為名的資料夾
        self.base_dir = os.getcwd()
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.folder_name = f"回後買上漲圖表(基礎版)_{self.today_str}"
        self.folder_path = os.path.join(self.base_dir, self.folder_name)
        
        # 建立資料夾
        os.makedirs(self.folder_path, exist_ok=True)

    def get_all_tickers(self):
        """從證交所與櫃買中心取得【上市+上櫃】所有股票代號與名稱"""
        tickers = []
        
        print(">> 正在取得【上市】股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            data = response.json()
            for item in data:
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
            print(f"   成功取得上市標的。")
        except Exception as e:
            pass 

        print(">> 正在取得【上櫃】股票清單...")
        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            data = response.json()
            for item in data:
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
            print(f"   成功取得上櫃標的。")
        except Exception as e:
            pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def draw_and_save_chart(self, df, ticker, stock_name, yesterday_high):
        """🌟 繪製 K 線圖、20MA 與突破線，並存成圖片"""
        chart_df = df.iloc[-125:].copy() # 取近 60 天資料
        
        # 基礎版只看 20MA
        apds = [
            mpf.make_addplot(chart_df['20MA'], color='deeppink', width=1.5)
        ]
        
        # 設定水平突破線 (藍色：昨日高點)
        hlines_config = dict(
            hlines=[yesterday_high],
            colors=['blue'],
            linestyle='-.',
            linewidths=1.5
        )
        
        # 處理檔名，避開特殊字元
        pure_code = ticker.split('.')[0]
        clean_stock_name = str(stock_name).replace("/", "").replace("\\", "").strip()
        filename = os.path.join(self.folder_path, f"{pure_code}_{clean_stock_name}.png")
        
        mc = mpf.make_marketcolors(up='r', down='g', edge='inherit', wick='inherit', volume='in')
        s  = mpf.make_mpf_style(marketcolors=mc, gridstyle=':', y_on_right=True)
        
        os.makedirs(self.folder_path, exist_ok=True)
        
        plot_df = chart_df.tail(125)
        hank_6m_df = plot_df.tail(125)
        mpf.plot(hank_6m_df, 
            type='candle', 
            volume=True, 
            addplot=apds,           
            hlines=hlines_config,   
            style=s,
            savefig=dict(fname=filename, dpi=100, bbox_inches='tight'),
            figratio=(16,9),  
            figscale=1.2,
            tight_layout=True
        )
        # 確保釋放記憶體
        plt.close('all')

    def run_scan(self, tickers, min_vol_lots=1000, chunk_size=50):
        """執行批次掃描 (朱家泓回後買上漲策略)"""
        results = []
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描與繪圖中...")
            
            try:
                # 🌟 確保 auto_adjust=True 還原權值
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True, auto_adjust=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < 25: continue 
                    
                    df['20MA'] = df['Close'].rolling(window=20).mean() 
                    df['5VMA'] = df['Volume'].rolling(window=5).mean() 
                    
                    today = df.iloc[-1]
                    yesterday = df.iloc[-2]
                    day_before = df.iloc[-3]
                    
                    is_above_20ma = today['Close'] > today['20MA']
                    is_pullback = (yesterday['Close'] < day_before['Close']) or (yesterday['Close'] < yesterday['Open'])
                    is_break_yesterday_high = today['Close'] > yesterday['High']
                    is_red_candle = today['Close'] > today['Open']
                    
                    avg_vol_5d = today['5VMA']
                    is_liquid = avg_vol_5d >= (min_vol_lots * 1000)
                    
                    if is_above_20ma and is_pullback and is_break_yesterday_high and is_red_candle and is_liquid:
                        vol_ratio = today['Volume'] / avg_vol_5d if avg_vol_5d > 0 else 0
                        
                        pure_code = ticker.split('.')[0]
                        yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                        stock_name = self.ticker_names.get(ticker, "未知") 
                        
                        # 🌟 觸發條件後，呼叫繪圖函數
                        self.draw_and_save_chart(df, ticker, stock_name, yesterday['High'])
                        
                        results.append({
                            '代號': ticker,
                            '股票名稱': stock_name,  
                            '今日收盤': round(today['Close'], 2),
                            '昨日高點': round(yesterday['High'], 2),
                            '今日成交量(張)': int(today['Volume'] / 1000),
                            '五日均量(張)': int(avg_vol_5d / 1000),
                            '量能放大倍數': round(vol_ratio, 1),
                            '月線位置(20MA)': round(today['20MA'], 2),
                            '雅虎股市連結': yahoo_link
                        })
                        
            except Exception as e:
                # 遇到錯誤略過，不中斷全市場掃描
                pass 
            
            time.sleep(1.2) 

        return pd.DataFrame(results)

def export_to_excel(df, filepath):
    """將結果導出為精美樣式的 Excel 表格"""
    if df.empty: return
    
    df = df.sort_values(by='量能放大倍數', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "掃描標的"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    
    fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid") 
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")  
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main
            cell.border = border_all
            
            col_name = headers[c_idx - 1]
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = align_center
            elif col_name in ['代號', '股票名稱']:
                cell.value = val
                cell.alignment = align_center
            elif col_name in ['今日收盤', '昨日高點', '月線位置(20MA)', '量能放大倍數']:
                cell.value = val
                cell.alignment = align_right
                if col_name != '量能放大倍數':
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0.0'
            elif col_name in ['今日成交量(張)', '五日均量(張)']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結':
                cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱':
            ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    
    wb.save(filepath)

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = TaiwanStockScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (朱家泓回後買：基礎版 + 自動繪圖)...\n")
        final_df = scanner.run_scan(stocks, min_vol_lots=1000)
        
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)
        
        print("\n" + "="*85)
        if not final_df.empty:
            print(f"🔥 掃描完成！共發現 {len(final_df)} 檔符合條件的標的。")
            print(f"📂 所有產出（Excel 與 K線圖）已完美收納至資料夾：\n👉 【 {scanner.folder_path} 】")
            print(f"   (圖表說明 -> 粉紅線: 20MA月線 | 藍虛線: 昨日高點突破線)")
            
            # 🌟 儲存 Excel 到專屬資料夾
            excel_filepath = os.path.join(scanner.folder_path, f"回後買上漲基礎版_{scanner.today_str}.xlsx")
            export_to_excel(final_df, excel_filepath)
        else:
            print("今日無符合條件標的。")
        print("="*85 + "\n")