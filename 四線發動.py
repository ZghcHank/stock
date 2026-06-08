
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    # 使用 getattr 避開字串 replace 關鍵字，完美斷開無限遞迴死結
    raw_download = getattr(yf, 'download')
    df = raw_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df








import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
import os
import matplotlib
matplotlib.use('Agg')  
import matplotlib.pyplot as plt 
import mplfinance as mpf
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class MA4BreakoutScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 
        
        # 絕對路徑設定，確保完美收納
        self.base_dir = os.getcwd()
        self.today_str = datetime.now().strftime("%Y%m%d")
        self.folder_name = f"四均線起漲圖表_{self.today_str}"
        self.folder_path = os.path.join(self.base_dir, self.folder_name)
        
        os.makedirs(self.folder_path, exist_ok=True)

    def get_all_tickers(self):
        tickers = []
        print(">> 正在取得全市場股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
        except Exception: pass 

        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            for item in response.json():
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
        except Exception: pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def draw_and_save_chart(self, df, ticker, stock_name):
        """繪製 K 線圖與 4 條均線，檢視糾結與突破姿態"""
        # 抓取近 80 天資料，讓前面的「盤整糾結區」能看得很清楚
        chart_df = df.iloc[-125:].copy() 
        
        # 設定四條均線的顏色與粗細
        apds = [
            mpf.make_addplot(chart_df['5MA'], color='darkorange', width=1.0),
            mpf.make_addplot(chart_df['10MA'], color='purple', width=1.0),
            mpf.make_addplot(chart_df['20MA'], color='deeppink', width=1.2),
            mpf.make_addplot(chart_df['60MA'], color='dodgerblue', width=1.5)
        ]
        
        pure_code = ticker.split('.')[0]
        clean_stock_name = str(stock_name).replace("/", "").replace("\\", "").strip()
        filename = os.path.join(self.folder_path, f"{pure_code}_{clean_stock_name}.png")
        
        mc = mpf.make_marketcolors(up='r', down='g', edge='inherit', wick='inherit', volume='in')
        s  = mpf.make_mpf_style(marketcolors=mc, gridstyle=':', y_on_right=True)
        
        os.makedirs(self.folder_path, exist_ok=True)
        
        plot_df = chart_df.tail(125)
        hank_6m_df = plot_df.tail(125)
        mpf.plot(hank_6m_df, 
            type='candle', 
            volume=True, 
            addplot=apds,           
            style=s,
            savefig=dict(fname=filename, dpi=100, bbox_inches='tight'),
            figratio=(16,9),  
            figscale=1.2,
            tight_layout=True
        )
        plt.close('all')

    def run_scan(self, tickers, max_tangle_pct=0.04, vol_ratio=2.0, min_vol_lots=1000, chunk_size=50):
        """
        執行掃描: 4條均線糾結 + 今日實體大紅K突破所有均線 + 爆量
        """
        results = []
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描與繪圖中...")
            
            try:
                # 抓取 6 個月資料以確保 60MA 運算正確
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True, auto_adjust=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    if len(df) < 65: continue 
                    
                    # 計算四大均線
                    df['5MA'] = df['Close'].rolling(window=5).mean()
                    df['10MA'] = df['Close'].rolling(window=10).mean()
                    df['20MA'] = df['Close'].rolling(window=20).mean()
                    df['60MA'] = df['Close'].rolling(window=60).mean()
                    df['5VMA'] = df['Volume'].rolling(window=5).mean()
                    
                    today = df.iloc[-1]
                    yesterday = df.iloc[-2]
                    
                    # 抓取「今日」與「昨日」的四條均線數值
                    ma_values_today = [today['5MA'], today['10MA'], today['20MA'], today['60MA']]
                    max_ma_today = max(ma_values_today)
                    
                    ma_values_yest = [yesterday['5MA'], yesterday['10MA'], yesterday['20MA'], yesterday['60MA']]
                    max_ma_yest = max(ma_values_yest)
                    min_ma_yest = min(ma_values_yest)
                    
                    # 1. 均線糾結度判定 (以昨日為基準，四條均線黏在一起，誤差在 4% 內)
                    tangle_pct = (max_ma_yest - min_ma_yest) / min_ma_yest
                    is_tangled = tangle_pct <= max_tangle_pct
                    
                    # 2. 剛脫離均線 (昨日收盤還在均線群內/下，今日強勢突破所有均線)
                    is_breakout = (yesterday['Close'] <= max_ma_yest * 1.01) and (today['Close'] > max_ma_today)
                    
                    # 3. 實體紅K表態 (收盤 > 開盤)
                    is_red_candle = today['Close'] > today['Open']
                    
                    # 4. 爆量攻擊 (今日量大於五日均量設定倍數)
                    avg_vol_5d = today['5VMA']
                    is_volume_burst = today['Volume'] > (avg_vol_5d * vol_ratio)
                    is_liquid = avg_vol_5d >= (min_vol_lots * 1000)
                    
                    if is_tangled and is_breakout and is_red_candle and is_volume_burst and is_liquid:
                        
                        stock_name = self.ticker_names.get(ticker, "未知") 
                        pure_code = ticker.split('.')[0]
                        yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                        actual_vol_ratio = today['Volume'] / avg_vol_5d if avg_vol_5d > 0 else 0
                        
                        # 呼叫畫圖
                        self.draw_and_save_chart(df, ticker, stock_name)
                        
                        results.append({
                            '代號': ticker,
                            '股票名稱': stock_name,
                            '今日收盤': round(today['Close'], 2),
                            '均線糾結度': f"{round(tangle_pct * 100, 2)}%",
                            '今日成交量(張)': int(today['Volume'] / 1000),
                            '五日均量(張)': int(avg_vol_5d / 1000),
                            '量能放大倍數': round(actual_vol_ratio, 1),
                            '雅虎股市連結': yahoo_link
                        })
                        
            except Exception as e:
                # print(f"\\n[發生隱藏錯誤] {e}\\n") 
                pass
            
            time.sleep(1.2) 

        return pd.DataFrame(results)

def export_to_excel(df, filepath):
    """將結果導出為精美樣式的 Excel 表格"""
    if df.empty: return
    
    # 依照均線糾結度排序，越緊密的排在越上面 (爆發力通常越強)
    df = df.sort_values(by='均線糾結度', ascending=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "四均線起漲精選"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    font_highlight = Font(name="微軟正黑體", size=11, bold=True, color="FF0000") 
    
    fill_header = PatternFill(start_color="800080", end_color="800080", fill_type="solid") # 紫色表頭代表強勢動能
    fill_zebra = PatternFill(start_color="F9F2F9", end_color="F9F2F9", fill_type="solid")  
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main
            cell.border = border_all
            
            col_name = headers[c_idx - 1]
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = align_center
            elif col_name in ['代號', '股票名稱']:
                cell.value = val
                cell.alignment = align_center
            elif col_name == '均線糾結度':
                cell.value = val
                cell.alignment = align_center
                cell.font = font_highlight 
            elif col_name in ['今日收盤', '量能放大倍數']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0.00' if col_name == '今日收盤' else '#,##0.0'
            elif col_name in ['今日成交量(張)', '五日均量(張)']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結':
                cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱':
            ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    
    wb.save(filepath)

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = MA4BreakoutScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (四均線糾結 + 帶量紅K起漲突破)...\n")
        
        # max_tangle_pct=0.04 代表四條均線的最高與最低差距必須小於 4% (極度糾結)
        # vol_ratio=2.0 代表今日成交量必須大於 5日均量的 2 倍 (明顯表態)
        final_df = scanner.run_scan(stocks, max_tangle_pct=0.04, vol_ratio=2.0, min_vol_lots=1000)
        
        print("\n" + "="*95)
        if not final_df.empty:
            print(f"🚀 掃描完成！共發現 {len(final_df)} 檔剛脫離【四均線糾結】的強勢起漲股。")
            print(f"📂 所有產出（紫色 Excel 與 4均線K線圖）已完美收納至資料夾：\n👉 【 {scanner.folder_path} 】")
            print(f"   (圖表均線說明 -> 橘:5MA | 紫:10MA | 粉:20MA | 藍:60MA)")
            
            excel_filepath = os.path.join(scanner.folder_path, f"四均線起漲精選_{scanner.today_str}.xlsx")
            export_to_excel(final_df, excel_filepath)
        else:
            print("❌ 今日無符合「四均線緊密糾結」且剛好「今日帶量突破」的標的。")
        print("="*95 + "\n")