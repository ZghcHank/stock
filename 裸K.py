
def cached_yf_download(*args, **kwargs):
    import os
    import pandas as pd
    import yfinance as yf
    from datetime import datetime
    
    ticker = args[0] if args else kwargs.get('tickers')
    period = kwargs.get('period', '1y')
    
    if ticker and isinstance(ticker, str):
        date_str = datetime.today().strftime("%Y%m%d")
        cache_file = os.path.join("yf_cache", date_str, f"{ticker}_{period}.pkl")
        fallback_file = os.path.join("yf_cache", date_str, f"{ticker}_1y.pkl")
        
        for f_path in [cache_file, fallback_file]:
            if os.path.exists(f_path):
                try:
                    return pd.read_pickle(f_path)
                except Exception:
                    pass
                    
    # 🌟 使用 getattr 避開字串 replace 關鍵字，完美斷開無限遞迴死結
    raw_download = getattr(yf, 'download')
    df = raw_download(*args, **kwargs)
    if ticker and isinstance(ticker, str) and not df.empty:
        try:
            date_str = datetime.today().strftime("%Y%m%d")
            cache_dir = os.path.join("yf_cache", date_str)
            os.makedirs(cache_dir, exist_ok=True)
            df.to_pickle(os.path.join(cache_dir, f"{ticker}_{period}.pkl"))
        except Exception:
            pass
    return df





import requests
if 'session' not in locals() and 'session' not in globals():
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Language': 'zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7'
    })

import yfinance as yf
import pandas as pd
import requests
import time
import urllib3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 關閉 SSL 驗證警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class LaoYuNakedKScanner:
    def __init__(self):
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        self.ticker_names = {} 

    def get_all_tickers(self):
        """從證交所與櫃買中心取得【上市+上櫃】所有股票代號與名稱"""
        tickers = []
        
        print(">> 正在取得【上市】股票清單...")
        try:
            tw_url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            response = requests.get(tw_url, headers=self.headers, verify=False)
            data = response.json()
            for item in data:
                code = item['Code']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TW"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['Name'] 
            print(f"   成功取得上市標的。")
        except Exception as e:
            pass 

        print(">> 正在取得【上櫃】股票清單...")
        try:
            two_url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            response = requests.get(two_url, headers=self.headers, verify=False)
            data = response.json()
            for item in data:
                code = item['SecuritiesCompanyCode']
                if len(code) == 4 and code.isdigit():
                    ticker = f"{code}.TWO"
                    tickers.append(ticker)
                    self.ticker_names[ticker] = item['CompanyName']
            print(f"   成功取得上櫃標的。")
        except Exception as e:
            pass

        print(f">> ✅ 總共集結 {len(tickers)} 檔全市場股票準備掃描。\n")
        return tickers

    def run_scan(self, tickers, support_lookback=40, trap_window=5, min_rr_ratio=2.0, min_vol_lots=500, chunk_size=50):
        """
        執行批次掃描 (老余裸K破底翻 + 賺賠比)
        """
        results = []
        total = len(tickers)

        for i in range(0, total, chunk_size):
            chunk = tickers[i:i + chunk_size]
            print(f"進度: {min(i + len(chunk), total)}/{total} 掃描中...")
            
            try:
                # 抓取近 3 個月資料已足夠計算
                df_all = cached_yf_download(" ".join(chunk), period="1y", group_by='ticker', progress=False, threads=True)
                
                for ticker in chunk:
                    if ticker not in df_all.columns.levels[0]: continue
                    
                    df = df_all[ticker].dropna()
                    # 確保資料長度大於觀察期
                    if len(df) < (support_lookback + trap_window + 5): continue 
                    
                    # 1. 劃分時間區塊
                    # 歷史區 (尋找支撐與壓力)
                    history_zone = df.iloc[-(support_lookback + trap_window) : -trap_window]
                    # 近期誘空區 (尋找破底的最低點)
                    trap_zone = df.iloc[-trap_window : -1]
                    today = df.iloc[-1]
                    
                    # 2. 定義關鍵價格 (畫線)
                    support_price = history_zone['Low'].min()      # 頸線支撐
                    resistance_price = history_zone['High'].max()  # 前波壓力 (目標價)
                    trap_low = trap_zone['Low'].min()              # 破底坑的最低點 (停損點)
                    
                    entry_price = today['Close']                   # 今日收盤價 (進場點)
                    
                    # 3. 型態條件判斷
                    # A. 確實有破底
                    is_breakdown = trap_low < support_price
                    # B. 今日強勢站回支撐
                    is_reclaim = entry_price > support_price
                    # C. 實體紅K表態 (不買留長上影線的假跌破)
                    is_red_candle = entry_price > today['Open']
                    # D. 基本流動性 (避開一天成交不到500張的殭屍股，以免滑價)
                    is_liquid = today['Volume'] >= (min_vol_lots * 1000)
                    
                    if is_breakdown and is_reclaim and is_red_candle and is_liquid:
                        
                        # 4. 核心：計算風險與報酬 (Risk-Reward)
                        risk = entry_price - trap_low          # 潛在虧損
                        reward = resistance_price - entry_price # 潛在獲利
                        
                        # 避免除以零或無獲利空間的極端情況
                        if risk > 0 and reward > 0:
                            rr_ratio = reward / risk
                            
                            # E. 賺賠比濾網 (大於設定的 2.0 倍)
                            if rr_ratio >= min_rr_ratio:
                                stock_name = self.ticker_names.get(ticker, "未知") 
                                pure_code = ticker.split('.')[0]
                                yahoo_link = f"https://tw.stock.yahoo.com/quote/{pure_code}"
                                
                                results.append({
                                    '代號': ticker,
                                    '股票名稱': stock_name,
                                    '賺賠比(RR)': round(rr_ratio, 2),
                                    '進場價(今日收盤)': round(entry_price, 2),
                                    '停損價(破底低點)': round(trap_low, 2),
                                    '目標價(前波壓力)': round(resistance_price, 2),
                                    '風險空間(元)': round(risk, 2),
                                    '今日成交量(張)': int(today['Volume'] / 1000),
                                    '雅虎股市連結': yahoo_link
                                })
                        
            except Exception as e:
                pass 
            
            time.sleep(1.2) 

        return pd.DataFrame(results)

def export_to_excel(df, filename="老余裸K_賺賠比精選.xlsx"):
    """將結果導出為精美樣式的 Excel 表格"""
    if df.empty:
        print(f">> 殘酷的市場：今日無符合賺賠比 >= 2.0 的破底翻標的。")
        return
    
    # 老余流的核心：依照「賺賠比」由大到小排序，尋找最具 CP 值的交易
    df = df.sort_values(by='賺賠比(RR)', ascending=False)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "裸K賺賠比精選"
    
    font_main = Font(name="微軟正黑體", size=11)
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_link = Font(name="微軟正黑體", size=11, color="0000FF", underline="single")
    font_highlight = Font(name="微軟正黑體", size=11, bold=True, color="008000") # 綠色高光
    
    fill_header = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid") # 墨綠色表頭
    fill_zebra = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")  
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    headers = list(df.columns)
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    
    for r_idx, row_data in enumerate(df.values, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = font_main
            cell.border = border_all
            
            col_name = headers[c_idx - 1]
            if col_name == '雅虎股市連結':
                cell.value = "點我查看"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = align_center
            elif col_name in ['代號', '股票名稱']:
                cell.value = val
                cell.alignment = align_center
            elif col_name == '賺賠比(RR)':
                cell.value = f"1 : {val}" # 格式化為 1 : X
                cell.alignment = align_center
                cell.font = font_highlight # 賺賠比加粗變綠色
            elif col_name in ['進場價(今日收盤)', '停損價(破底低點)', '目標價(前波壓力)', '風險空間(元)']:
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0.00'
            elif col_name == '今日成交量(張)':
                cell.value = val
                cell.alignment = align_right
                cell.number_format = '#,##0'
                
            if r_idx % 2 == 0 and col_name != '雅虎股市連結':
                cell.fill = fill_zebra
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        if headers[col[0].column - 1] == '股票名稱':
            ws.column_dimensions[col_letter].width = max(max_len * 2, 14) 
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
        
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2" 
    
    wb.save(filename)
    print(f">> ✅ 成功產出 Excel：{filename}")

# ================= 執行區 =================
if __name__ == "__main__":
    scanner = LaoYuNakedKScanner()
    stocks = scanner.get_all_tickers()
    
    if stocks:
        print("開始執行全市場掃描 (老余流：裸K破底翻 + 賺賠比 >= 2)...\n")
        
        # 執行掃描：尋找賺賠比至少 1:2 的標的
        final_df = scanner.run_scan(stocks, min_rr_ratio=2.0)

# 💡 加上日期檔名
        today_str = datetime.now().strftime("%Y%m%d")
        dynamic_filename = f"老余裸K_賺賠比精選_{today_str}.xlsx"
        
        export_to_excel(final_df, filename=dynamic_filename)
        
        pd.set_option('display.unicode.ambiguous_as_wide', True)
        pd.set_option('display.unicode.east_asian_width', True)
        
        print("\n" + "="*90)
        if not final_df.empty:
            print(f"🔥 掃描完成！發現以下【高性價比】標的 (已依賺賠比排序)：\n")
            print(final_df.drop(columns=['雅虎股市連結']).to_string(index=False))
        else:
            print("❌ 殘酷的市場：今日無符合「賺賠比 >= 2.0」的破底翻標的。不勉強進場！")
        print("="*90 + "\n")
            
        export_to_excel(final_df)